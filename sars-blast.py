import requests
import os
import shutil
import urllib.parse
import urllib.request
import openpyxl
from pathlib import Path
import time
import re
import os

path_to_spreadsheet = input("What is the name of the spreadsheet that you would like to use (must be in the same folder as this program): ")

if "output-data" in os.listdir():
    shutil.rmtree('output-data')
    time.sleep(1)

os.mkdir("output-data")


xlsx_file = Path(path_to_spreadsheet)
x = openpyxl.load_workbook(xlsx_file)
sheet = x.active

def find_all(lines, tag):
    return_list = []
    for i in lines:
        if tag in i:
            return_list.append(i)
    return return_list

GENES = []
TRANSFER_GENES = []
HUMAN_GENES = []
GO_TERMS = []
ENTRY_NAME = []
FASTA = []

ii = 0

#441 rows, do in bits of 100

print()

START = input("What entry should it start from, or press enter to start from the beginning: ")
END = input("How many entries should I run, or press enter to do all: ")

if START.isnumeric() == False:
    START = 0
else:
    START = int(START)

if END.isnumeric() == False:
    END = START + 1000000000000000
else:
    END = int(END)

length_of_sheet = [0 for _ in sheet.iter_rows()]

print()

for row in sheet.iter_rows():
    
    if row[1].value != sheet["B1"].value and START <= ii <= END:
        GENES.append({"column": row[1].column, "row": row[1].row, "value": row[1].value})

    ii+=1

url_forOrtho = 'https://www.uniprot.org/uniref/'
url_forHuman = 'https://www.uniprot.org/uploadlists/'
url_forGoTerms = 'https://www.uniprot.org/uniprot/'
url_forEntry = 'https://www.uniprot.org/uniprot/'

#save every 10 in case anything goes wrong
def process_gene(gene, accumulator):

    if accumulator % 10 == 0:
        x.save('./{}'.format(path_to_spreadsheet))
        print("\t[auto-saving excel sheet]\n")
        time.sleep(4)

    COL, ROW, VAL = tuple(gene.values())

    print("STARTING ENTRY NUMBER #" + str(ROW-1) + "...")

    GENE = VAL + ' taxonomy: "Homo sapiens (Human) [9606]"'

    params_forOrtho = {
        'sort': 'score',
        'query': GENE,
        'format': 'tab'
    }

    params_forEntry = {
        'query': "accession:"+VAL,
        'format': 'tab'
    }

    data = urllib.parse.urlencode(params_forOrtho)
    data = data.encode('utf-8')
    req = urllib.request.Request(url_forOrtho, data)
    
    try:
        with urllib.request.urlopen(req) as f:
            response = f.read()
    except:
        time.sleep(5)
        with urllib.request.urlopen(req) as f:
            response = f.read()

    transfer_response = response.decode('utf-8')

    if len(transfer_response) > 0:
        trans = transfer_response.split("\n")[1].split("\t")[0]
    else:
        trans= "N/A"
    
    data = urllib.parse.urlencode(params_forEntry).encode('utf-8')
    req = urllib.request.Request(url_forEntry, data)
    
    try:
        with urllib.request.urlopen(req) as f:
            response = f.read()
    except:
        time.sleep(5)
        with urllib.request.urlopen(req) as f:
            response = f.read()

    entry_name = response.decode('utf-8')

    ENTRY_NAME = entry_name.split("\n")[1].split('\t')[1]


    fasta_url = 'https://uniprot.org/uniprot/{}.fasta'.format(VAL)
    fasta_response = requests.get(fasta_url).text

    if len(fasta_response) > 0:
        os.mkdir('output-data/{}'.format(VAL))
        f = open('output-data/{}/{}.fasta'.format(VAL, VAL), 'w')
        f.write("".join(fasta_response))
        f.close()

    GO_TERMS = []

    print("SAVING ENTRY NUMBER #" + str(ROW-1) + "...")

    if trans != "N/A":
        if "UniRef100" in trans:
            from_forHuman = "NF100"
        elif "UniRef90" in trans:
            from_forHuman = "NF90"
        elif "UniRef50" in trans:
            from_forHuman = "NF50"
        else:
            print("THIS IS NOT A THING")
        
        done = False

        while not done:

            params_forHuman = {
                'from': from_forHuman,
                'to': 'ID',
                'format': 'tab',
                'query': trans
            }

            data = urllib.parse.urlencode(params_forHuman)
            data = data.encode('utf-8')
            req = urllib.request.Request(url_forHuman, data)
            try:
                with urllib.request.urlopen(req) as f:
                    response = f.read()
            except:
                time.sleep(5)
                with urllib.request.urlopen(req) as f:
                    response = f.read()

            human_response = response.decode('utf-8')

            if "_HUMAN" in human_response:
                HUMAN_GENES_TEMP = find_all('\t'.join(human_response.split('\n')).split('\t'), "_HUMAN")
                HUMAN_GENES = HUMAN_GENES_TEMP

                j = 0
                
                for gene in HUMAN_GENES_TEMP:

                    url = 'https://uniprot.org/uniprot/{}.txt'.format(gene)

                    go_terms_response = requests.get(url).text

                    if len(go_terms_response) > 0 and " GO:" in go_terms_response:
                        go_terms = re.sub(r'(DR   GO; GO:)[0-9]{7}(; )', '', '\n'.join(find_all('\t'.join(go_terms_response.split("\n")).split('\t'), "GO")))
                        try:
                            go_terms = [i[:i.index(';')] for i in go_terms.split('\n') if i[:2] in "C:P:F:"]
                        except:
                            go_terms = go_terms.split('\n')

                        go_terms.insert(0, "Gene {}: \n\t".format(gene))
                        go_terms = " ".join(go_terms)
                        GO_TERMS.append(go_terms)
                    else:
                        GO_TERMS.append("Gene {}: \n\tNO GO TERMS ASSOCIATED".format(gene))

                    j+=1
                f = open("output-data/"+VAL+'/GO_terms.txt', 'w')
                f.write("\n\n".join(GO_TERMS))
                f.close()

                done = True

            else:

                done = False

                if from_forHuman == "NF100":
                    from_forHuman = "NF90"
                    trans = "UniRef90_" + trans.split("_")[1]
                elif from_forHuman == "NF90":
                    from_forHuman = "NF50"
                    trans = "UniRef50_" + trans.split("_")[1]
                else:
                    done = True
                    
                    GO_TERMS = "N/A"
                    HUMAN_GENES = 'N/A'
    else:
        HUMAN_GENES = "N/A"
        GO_TERMS = 'N/A'

    sheet.cell(column= COL+1, row=ROW).value = str(ENTRY_NAME)
    sheet.cell(column= COL+3, row=ROW).value = str(HUMAN_GENES) if str(HUMAN_GENES) == "N/A" else ', '.join(HUMAN_GENES)
    sheet.cell(column= COL+7, row=ROW).value = '=HYPERLINK(\"./output-data/{}/GO_terms.txt\",\"Link to GO terms\")'.format(VAL) if str(GO_TERMS) != "N/A" else "N/A"
    sheet.cell(column= COL+8, row=ROW).value = '=HYPERLINK(\"./output-data/{}/\",\"Link to FASTA\")'.format(VAL)

    print("FINISHED ENTRY NUMBER #" + str(ROW-1) + "!\n")



for i in range(len(length_of_sheet)-1):
    process_gene(GENES.pop(0), i+1)

x.save('./{}'.format(path_to_spreadsheet))
os.system('node javascript-scripts/main.js --path={}'.format(path_to_spreadsheet))